# excel_backend.py

from openpyxl import load_workbook
from functools import lru_cache
import os

EXCEL_FILE = "WWY_ProbioticDrink_Model_v1_DASHBOARD.xlsx"


@lru_cache(maxsize=1)
def _load_wb():
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    return load_workbook(EXCEL_FILE, data_only=True)


def get_fruit_master():
    """Return list of fruits with sugar, sweetness, tartness from FruitMaster sheet."""
    wb = _load_wb()
    ws = wb["FruitMaster"]
    fruits = []

    # Assuming headers in row 1: Fruit | Sugar_g_per_100ml | Sweetness_Score_1to10 | Tartness_Score_1to10 | Notes
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, sugar, sweet, tart, notes = row
        if not name:
            continue
        fruits.append({
            "name": str(name),
            "sugar": float(sugar or 0),
            "sweet": int(sweet or 0),
            "tart": int(tart or 0),
            "notes": notes or "",
        })
    return fruits


def get_cost_table():
    """Read Costing sheet into a dict keyed by ingredient (lowercase)."""
    wb = _load_wb()
    if "Costing" not in wb.sheetnames:
        return {}
    ws = wb["Costing"]
    costs = {}
    # Assuming: Ingredient | Cost_per_L_or_kg | Usage_Unit | Cost_for_Batch
    for row in ws.iter_rows(min_row=2, values_only=True):
        ingredient, cost_per_unit, unit, _ = row
        if not ingredient:
            continue
        key = str(ingredient).strip().lower()
        costs[key] = {
            "ingredient": ingredient,
            "cost_per_unit": float(cost_per_unit or 0),
            "unit": unit or "",
        }
    return costs


def get_co2_safety_table():
    """Read CO2Safety sheet into a list of rows."""
    wb = _load_wb()
    if "CO2Safety" not in wb.sheetnames:
        return []
    ws = wb["CO2Safety"]
    rows = []
    # Sugar_g_per_L | Temp_C | Max_Time_Hours | Risk
    for row in ws.iter_rows(min_row=2, values_only=True):
        sugar, temp, hours, risk = row
        if sugar is None or temp is None:
            continue
        rows.append({
            "sugar_g_L": float(sugar),
            "temp_C": float(temp),
            "max_hours": float(hours or 0),
            "risk": str(risk or ""),
        })
    return rows


def _style_aliases():
    return {
        "tropical": ["tropical", "mango", "pineapple", "aromatic"],
        "berry": ["berry", "red", "grape"],
        "citrus": ["citrus", "acid", "bright"],
        "neutral": ["neutral", "refreshing", "balanced"],
    }


def auto_suggest_from_excel(target_sweet, target_tart, style, total_juice_ml_per_L=80.0, batch_l=3.0, temp_C=28.0):
    """Use FruitMaster to suggest 4 fruits that best match target sweet/tart."""
    fruits = get_fruit_master()
    style = (style or "").lower().strip()
    style_tags = _style_aliases().get(style, [])

    scored = []
    for f in fruits:
        base_distance = abs(f["sweet"] - target_sweet) + abs(f["tart"] - target_tart)
        style_bonus = 0
        if style_tags and any(tag in (f["notes"] or "").lower() for tag in style_tags):
            style_bonus = -1
        score = base_distance + style_bonus
        scored.append((score, f))

    scored.sort(key=lambda x: x[0])
    top_fruits = [f for _, f in scored[:4]]

    pct_pattern = [0.4, 0.3, 0.2, 0.1]
    fruits_out = []
    sugar_total_g_L = 0.0

    for i, f in enumerate(top_fruits):
        pct = pct_pattern[i] if i < len(pct_pattern) else 0.0
        juice_ml_per_L = total_juice_ml_per_L * pct
        sugar_g_L = f["sugar"] * juice_ml_per_L / 100.0
        sugar_total_g_L += sugar_g_L

        fruits_out.append({
            "name": f["name"],
            "pct": pct,
            "juice_ml_per_L": round(juice_ml_per_L, 2),
        })

    co2_vols = sugar_total_g_L * 0.24
    abv_percent = sugar_total_g_L * 0.065
    safety_flag = "OK (≤ 8 g/L)" if sugar_total_g_L <= 8 else "Too high – reduce juice/sugar"

    # Add complete formulation breakdown
    formulation = _calculate_formulation(batch_l, total_juice_ml_per_L)

    # Calculate optimal fermentation time
    ferment_time = _calculate_optimal_ferment_time(sugar_total_g_L, temp_C)

    return {
        "fruits": fruits_out,
        "sugar_g_per_L": round(sugar_total_g_L, 2),
        "co2_vols": round(co2_vols, 2),
        "abv_percent": round(abv_percent, 3),
        "safety_flag": safety_flag,
        "formulation": formulation,
        "ferment_time": ferment_time,
        "temp_C": temp_C,
    }


def _lookup_fruit_sugar(name):
    name = (name or "").strip().lower()
    for f in get_fruit_master():
        if f["name"].strip().lower() == name:
            return f["sugar"]
    return 0.0


def _estimate_cost_for_blend(fruits, batch_l):
    """Estimate cost for a given blend using Costing sheet."""
    costs = get_cost_table()
    total_cost = 0.0

    for f in fruits:
        fruit_name = f["name"]
        juice_ml_batch = f["juice_ml_batch"]
        key_candidates = [
            fruit_name.strip().lower(),
            (fruit_name + " juice").strip().lower(),
        ]
        cost_info = None
        for key in key_candidates:
            if key in costs:
                cost_info = costs[key]
                break

        if cost_info:
            if "per L" in cost_info["unit"]:
                liters = juice_ml_batch / 1000.0
                total_cost += liters * cost_info["cost_per_unit"]
            elif "per kg" in cost_info["unit"]:
                # rough approximation: assume juice density ~1 kg/L
                liters = juice_ml_batch / 1000.0
                total_cost += liters * cost_info["cost_per_unit"]
            # else: ignore or extend logic as needed
    return round(total_cost, 2)


def _lookup_safety_row(sugar_g_L, temp_C):
    """Find the closest safety row given sugar and temp."""
    rows = get_co2_safety_table()
    if not rows:
        return None
    # Simple: choose row where sugar_g_L <= sheet sugar and closest temp
    candidates = [r for r in rows if sugar_g_L <= r["sugar_g_L"] + 0.1]
    if not candidates:
        candidates = rows
    candidates.sort(key=lambda r: (abs(r["temp_C"] - temp_C), r["sugar_g_L"]))
    return candidates[0]


def _calculate_optimal_juice_amount(fruit_names, target_sugar_g_L=7.0):
    """
    Calculate optimal juice amount per liter to achieve target sugar level.

    Args:
        fruit_names: List of fruit names (can include empty strings)
        target_sugar_g_L: Target sugar content in g/L (default 7.0 for safety)

    Returns:
        dict with recommended juice amounts and reasoning
    """
    # Filter out empty fruit names
    selected_fruits = [name for name in fruit_names if name and name.strip()]

    if not selected_fruits:
        # Default if no fruits selected
        return {
            "recommended_ml_per_L": 80,
            "intensity": "medium",
            "reasoning": "Default recommendation: 80 ml/L for balanced flavor",
            "sugar_estimate_g_L": 7.0
        }

    # Get fruit data and calculate average sugar content
    total_sugar = 0
    fruit_count = 0

    for name in selected_fruits:
        sugar = _lookup_fruit_sugar(name)
        if sugar > 0:
            total_sugar += sugar
            fruit_count += 1

    if fruit_count == 0:
        avg_sugar = 10.0  # Default assumption
    else:
        avg_sugar = total_sugar / fruit_count

    # Calculate juice amount to achieve target sugar
    # Formula: sugar_g_L = (avg_sugar_g_per_100ml * juice_ml_per_L) / 100
    # Therefore: juice_ml_per_L = (target_sugar_g_L * 100) / avg_sugar_g_per_100ml

    optimal_juice_ml_L = (target_sugar_g_L * 100) / avg_sugar if avg_sugar > 0 else 80

    # Round to nearest 5 ml for practicality
    optimal_juice_ml_L = round(optimal_juice_ml_L / 5) * 5

    # Clamp to reasonable range (40-150 ml/L)
    optimal_juice_ml_L = max(40, min(150, optimal_juice_ml_L))

    # Calculate actual sugar with this amount
    actual_sugar_g_L = (avg_sugar * optimal_juice_ml_L) / 100

    # Determine intensity
    if optimal_juice_ml_L < 60:
        intensity = "light"
    elif optimal_juice_ml_L <= 90:
        intensity = "medium"
    else:
        intensity = "strong"

    # Generate reasoning
    fruit_list = ", ".join(selected_fruits[:3])
    if len(selected_fruits) > 3:
        fruit_list += f" and {len(selected_fruits) - 3} more"

    reasoning = (
        f"Based on {fruit_list} (avg {avg_sugar:.1f}g sugar/100ml), "
        f"{optimal_juice_ml_L} ml/L will give ~{actual_sugar_g_L:.1f}g/L sugar - "
        f"safe and balanced"
    )

    return {
        "recommended_ml_per_L": int(optimal_juice_ml_L),
        "intensity": intensity,
        "reasoning": reasoning,
        "sugar_estimate_g_L": round(actual_sugar_g_L, 1),
        "avg_fruit_sugar": round(avg_sugar, 1)
    }


def _calculate_optimal_ferment_time(sugar_g_L, temp_C, humidity=60):
    """
    Calculate optimal fermentation time based on temperature, sugar, and humidity.

    Args:
        sugar_g_L: Sugar content in g/L
        temp_C: Temperature in Celsius
        humidity: Relative humidity percentage

    Returns:
        dict with fermentation time recommendations
    """
    # Base fermentation time calculation (in hours)
    # Lower sugar = faster fermentation
    # Higher temperature = faster fermentation
    # Optimal range: 20-30°C

    # Base time at 25°C with 8 g/L sugar
    base_hours = 24

    # Temperature factor (optimal at 25-28°C)
    if temp_C < 20:
        temp_factor = 1.5  # Slower fermentation in cold
    elif temp_C <= 25:
        temp_factor = 1.2
    elif temp_C <= 30:
        temp_factor = 1.0  # Optimal range
    elif temp_C <= 35:
        temp_factor = 0.8  # Faster but riskier
    else:
        temp_factor = 0.6  # Very fast, high risk

    # Sugar factor (more sugar = more activity but also more time needed for balance)
    if sugar_g_L <= 4:
        sugar_factor = 0.7  # Less sugar, faster
    elif sugar_g_L <= 6:
        sugar_factor = 0.85
    elif sugar_g_L <= 8:
        sugar_factor = 1.0  # Optimal range
    else:
        sugar_factor = 1.2  # More sugar, needs more time

    # Calculate time range
    min_hours = round(base_hours * temp_factor * sugar_factor * 0.75)
    max_hours = round(base_hours * temp_factor * sugar_factor * 1.25)
    optimal_hours = round(base_hours * temp_factor * sugar_factor)

    # Determine fermentation phase recommendations
    if temp_C >= 25 and temp_C <= 30:
        recommendation = "Ideal conditions for fermentation"
        quality = "optimal"
    elif temp_C >= 20 and temp_C < 25:
        recommendation = "Good conditions, but fermentation will be slower"
        quality = "good"
    elif temp_C > 30 and temp_C <= 35:
        recommendation = "Warm conditions - watch closely to avoid over-carbonation"
        quality = "caution"
    elif temp_C < 20:
        recommendation = "Cool conditions - fermentation will be slow. Consider warming."
        quality = "slow"
    else:
        recommendation = "Extreme temperature - not recommended for fermentation"
        quality = "danger"

    return {
        "min_hours": min_hours,
        "max_hours": max_hours,
        "optimal_hours": optimal_hours,
        "recommendation": recommendation,
        "quality": quality,
        "phase_1_hours": round(optimal_hours * 0.6),  # Active fermentation
        "phase_2_hours": round(optimal_hours * 0.4),  # Settling/carbonation
    }


def _calculate_formulation(batch_l, juice_ml_per_L, ginger_bug_pct=0.1, lemon_ml_per_L=10):
    """
    Calculate complete formulation breakdown for the batch.

    Args:
        batch_l: Total batch size in liters
        juice_ml_per_L: Total fruit juice per liter
        ginger_bug_pct: Ginger bug as % of total batch (default 10%)
        lemon_ml_per_L: Lemon juice per liter (default 10ml)

    Returns:
        dict with ingredient breakdown
    """
    total_batch_ml = batch_l * 1000

    # Calculate each ingredient
    total_fruit_juice_ml = juice_ml_per_L * batch_l
    lemon_juice_ml = lemon_ml_per_L * batch_l
    ginger_bug_ml = total_batch_ml * ginger_bug_pct

    # Water makes up the rest
    water_ml = total_batch_ml - total_fruit_juice_ml - lemon_juice_ml - ginger_bug_ml

    return {
        "total_batch_ml": round(total_batch_ml, 1),
        "water_ml": round(water_ml, 1),
        "total_fruit_juice_ml": round(total_fruit_juice_ml, 1),
        "lemon_juice_ml": round(lemon_juice_ml, 1),
        "ginger_bug_ml": round(ginger_bug_ml, 1),
        "ginger_bug_pct": ginger_bug_pct * 100,
        "lemon_ml_per_L": lemon_ml_per_L,
    }


def calculate_blend_manual(fruit_names, pcts, juice_ml_per_L, batch_l, temp_C=28.0):
    """
    Manual mode:
      fruit_names: list of 4 names (can be empty)
      pcts: list of 4 floats (fractions, should sum ~1)
    """
    # Auto-correct percentages to sum to 100%
    original_pcts = pcts.copy()
    total_pct = sum(pcts)
    pct_corrected = False

    # Only normalize if we have non-zero percentages
    if total_pct > 0 and abs(total_pct - 1.0) > 0.001:
        # Normalize percentages to sum to 1.0
        pcts = [p / total_pct for p in pcts]
        pct_corrected = True

    fruits_out = []
    sugar_total_g_L = 0.0

    for name, pct, orig_pct in zip(fruit_names, pcts, original_pcts):
        if not name or pct <= 0:
            continue
        sugar_per_100 = _lookup_fruit_sugar(name)
        juice_ml_L = juice_ml_per_L * pct
        juice_ml_batch = juice_ml_L * batch_l
        sugar_g_L = sugar_per_100 * juice_ml_L / 100.0
        sugar_total_g_L += sugar_g_L

        fruit_data = {
            "name": name,
            "pct": pct,
            "juice_ml_per_L": round(juice_ml_L, 2),
            "juice_ml_batch": round(juice_ml_batch, 2),
            "sugar_g_L": round(sugar_g_L, 2),
        }

        # Include original percentage if it was corrected
        if pct_corrected:
            fruit_data["original_pct"] = orig_pct

        fruits_out.append(fruit_data)

    co2_vols = sugar_total_g_L * 0.24
    abv_percent = sugar_total_g_L * 0.065
    safety_flag = "OK (≤ 8 g/L)" if sugar_total_g_L <= 8 else "Too high – reduce juice/sugar"

    safety_row = _lookup_safety_row(sugar_total_g_L, temp_C)
    est_cost = _estimate_cost_for_blend(fruits_out, batch_l)

    # Add complete formulation breakdown
    formulation = _calculate_formulation(batch_l, juice_ml_per_L)

    # Calculate optimal fermentation time
    ferment_time = _calculate_optimal_ferment_time(sugar_total_g_L, temp_C)

    result = {
        "fruits": fruits_out,
        "sugar_g_per_L": round(sugar_total_g_L, 2),
        "co2_vols": round(co2_vols, 2),
        "abv_percent": round(abv_percent, 3),
        "safety_flag": safety_flag,
        "batch_l": batch_l,
        "juice_ml_per_L": juice_ml_per_L,
        "cost_estimate": est_cost,
        "formulation": formulation,
        "pct_corrected": pct_corrected,
        "ferment_time": ferment_time,
        "temp_C": temp_C,
    }

    # Add correction details if percentages were auto-corrected
    if pct_corrected:
        result["correction_message"] = f"Percentages auto-corrected from {round(total_pct * 100, 1)}% to 100%"
        result["original_total_pct"] = round(total_pct * 100, 1)

    if safety_row:
        result["safety_detail"] = {
            "temp_C": safety_row["temp_C"],
            "max_hours": safety_row["max_hours"],
            "risk": safety_row["risk"],
        }

    return result
